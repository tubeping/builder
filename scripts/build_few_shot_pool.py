"""인스타 공구 캡션 → Few-shot 풀 JSON 생성.

소스: tubeping-sourcing/output/reports/gonggu_instagram_MASTER_*.xlsx (최신 1개)
출력: tubeping_builder/lib/prompts/fewShotPool.json

규칙:
- group(상위 카테고리)별로 likes 상위 N개 추출 (기본 N=25)
- 캡션 길이 60~1200자만 (너무 짧거나 길면 학습에 부적합)
- '#광고', '#sp', '#유료광고' 등 광고 표시 제거
- 노이즈(반복 줄바꿈, 잡 해시태그) 정리
- 우리 카테고리 분류로 매핑 (group → uiCategory)

JSON 구조:
{
  "generatedAt": "2026-05-18T...",
  "sourceFile": "gonggu_instagram_MASTER_20260518_0012.xlsx",
  "totalSamples": 180,
  "pool": {
    "beauty":     [ {url, username, caption, likes, hashtag}, ... ],
    "food":       [ ... ],
    "pet":        [ ... ],
    "health":     [ ... ],
    "kitchen":    [ ... ],
    "kids":       [ ... ],
    "medical":    [ ... ],
    "other":      [ ... ]
  }
}
"""
from __future__ import annotations

import io
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent  # tubeping_builder/
SOURCING = ROOT.parent  # tubeping-sourcing/
REPORTS_DIR = SOURCING / "output" / "reports"
OUT_PATH = ROOT / "lib" / "prompts" / "fewShotPool.json"

# 카테고리당 샘플 수
PER_CATEGORY = 25
MIN_CAPTION_LEN = 60
MAX_CAPTION_LEN = 1200

# xlsx group → 우리 시스템 uiCategory 매핑
GROUP_TO_UI: dict[str, str] = {
    "뷰티": "beauty",
    "반려동물_전체": "pet",
    "반려동물": "pet",
    "식품": "food",
    "멜론_과일": "food",
    "올리브오일": "food",
    "헬스_다이어트": "health",
    "건강_다이어트": "health",
    "키친_리빙": "kitchen",
    "리빙": "kitchen",
    "키즈_베이비": "kids",
    "키즈": "kids",
    "의료_종사자": "medical",
    "의료": "medical",
}

AD_TAG_PATTERNS = [
    r"#유료광고[^\s]*",
    r"#광고[^\s]*",
    r"#sp(?:on)?[^\s]*",
    r"#ad[^\s]*",
    r"#협찬[^\s]*",
]


def find_latest_master() -> Path:
    """가장 최근 MASTER xlsx."""
    candidates = sorted(REPORTS_DIR.glob("gonggu_instagram_MASTER_*.xlsx"))
    if not candidates:
        raise FileNotFoundError(f"No MASTER xlsx in {REPORTS_DIR}")
    return candidates[-1]


def clean_caption(text: str) -> str:
    """캡션 노이즈 제거."""
    t = text or ""
    # 광고 표시 제거
    for pat in AD_TAG_PATTERNS:
        t = re.sub(pat, "", t, flags=re.IGNORECASE)
    # zero-width / 공백 통일
    t = t.replace("​", "").replace("⁣", "")
    # 연속 줄바꿈 ≥3 → 2로
    t = re.sub(r"\n{3,}", "\n\n", t)
    # 줄 끝 공백 제거
    t = "\n".join(line.rstrip() for line in t.splitlines())
    return t.strip()


def map_category(group: str) -> str:
    if not group:
        return "other"
    return GROUP_TO_UI.get(group.strip(), "other")


def to_int(v) -> int:
    try:
        return int(v)
    except Exception:
        return 0


def build_pool() -> dict:
    src = find_latest_master()
    print(f"source: {src.name}")
    wb = openpyxl.load_workbook(src, read_only=True)
    ws = wb["1_전체게시물"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(header, row))
        caption = clean_caption(d.get("caption") or "")
        if len(caption) < MIN_CAPTION_LEN or len(caption) > MAX_CAPTION_LEN:
            continue
        ui_cat = map_category(d.get("group") or "")
        rows.append(
            {
                "uiCategory": ui_cat,
                "subcat": (d.get("subcat") or "").strip(),
                "url": (d.get("url") or "").strip(),
                "username": (d.get("username") or "").strip(),
                "caption": caption,
                "likes": to_int(d.get("likes")),
                "hashtag": (d.get("hashtag") or "").strip(),
            }
        )
    print(f"raw rows passing filter: {len(rows)}")

    # 카테고리별 likes 내림차순 정렬 → 상위 PER_CATEGORY
    by_cat: dict[str, list] = {}
    for r in rows:
        by_cat.setdefault(r["uiCategory"], []).append(r)
    for cat in by_cat:
        by_cat[cat].sort(key=lambda r: r["likes"], reverse=True)
        # 중복 username 너무 많이 안 잡히도록 한 사람당 최대 3개
        seen: dict[str, int] = {}
        picked = []
        for r in by_cat[cat]:
            n = seen.get(r["username"], 0)
            if n >= 3:
                continue
            seen[r["username"]] = n + 1
            picked.append(r)
            if len(picked) >= PER_CATEGORY:
                break
        by_cat[cat] = picked

    total = sum(len(v) for v in by_cat.values())
    pool = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sourceFile": src.name,
        "totalSamples": total,
        "perCategoryCap": PER_CATEGORY,
        "pool": {cat: by_cat.get(cat, []) for cat in sorted(by_cat.keys())},
    }
    return pool


def main() -> int:
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    pool = build_pool()
    OUT_PATH.write_text(
        json.dumps(pool, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"wrote {OUT_PATH} (samples={pool['totalSamples']})")
    for cat, items in pool["pool"].items():
        print(f"  {cat}: {len(items)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
